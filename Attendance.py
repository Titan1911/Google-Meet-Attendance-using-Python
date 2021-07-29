# web driver imports
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
# excel file imports
from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
# time import
from time import sleep
# installing packages import
import os


# The function directly installs the required libraries
def install(package):
    os.system("pip install " + str(package))
    print("Installed", package.upper())


if __name__ == '__main__':
    install("openpyxl")
    install("selenium")
    gmail_id = input('Enter your gmail id: ')
    gmail_password = input('Enter your gmail password: ')
    meet_link = input('Paste the link of the Google Meet: ')

    # Some necessary things for automation with google driver
    opt = Options()
    opt.add_argument("--disable-infobars")
    opt.add_argument("start-maximized")
    opt.add_argument("--disable-extensions")

    # This part allows the notifications, mic and camera permissions
    # Pass the argument 1 to allow and 2 to block
    opt.add_experimental_option("prefs", {
        "profile.default_content_setting_values.media_stream_mic": 1,
        "profile.default_content_setting_values.media_stream_camera": 1,
        "profile.default_content_setting_values.geolocation": 1,
        "profile.default_content_setting_values.notifications": 1
      })

    # Sign in to google
    driver = webdriver.Chrome(options=opt, executable_path=r"Enter your path here") # eg. r"C:\Users\hp\Downloads\chromedriver_win32\chromedriver.exe"
    driver.get('https://stackoverflow.com/users/signup?ssrc=head&returnurl=%2fusers%2fstory%2fcurrent%27')  # signing in to google through stack overflow
    sleep(2)
    driver.find_element_by_xpath('//*[@id="openid-buttons"]/button[1]').click()  # signing in with google
    driver.find_element_by_xpath('//input[@type="email"]').send_keys(gmail_id)  # entering the gmail id
    driver.find_element_by_xpath('//*[@id="identifierNext"]').click()
    sleep(2)
    driver.find_element_by_xpath('//input[@type="password"]').send_keys(gmail_password)  # entering the password
    driver.find_element_by_xpath('//*[@id="passwordNext"]').click()
    sleep(2)

    driver.get('https://meet.google.com/')

    # Enter the meeting
    # Case when logged in with personal gmail account
    driver.find_element_by_css_selector('input#i3').send_keys(meet_link)  # Enter a code or link
    sleep(1)
    driver.find_element_by_css_selector('button.VfPpkd-LgbsSe.VfPpkd-LgbsSe-OWXEXe-dgl2Hf.ksBjEc.lKxP2d.cjtUbb').click()  # join
    sleep(2)
    cam_mic_selectors = driver.find_elements_by_css_selector('div.U26fgb.JRY2Pb.mUbCce.kpROve')  # camera and mic
    for e in cam_mic_selectors:
        e.click()

    sleep(2)
    driver.find_element_by_css_selector('div.uArJ5e.UQuaGc.Y5sE8d.uyXBBb.xKiqt').click()  # join now
    sleep(5)
    driver.find_element_by_xpath('//*[@id="ow3"]/div[1]/div/div[9]/div[3]/div[10]/div[3]/div[2]/div/div/div[2]/span/button/i[1]').click()  # participant list
    sleep(1)
    names = driver.find_elements_by_css_selector('span.ZjFb7c')  # participants
    for e in names[1:]:
        print(e.text)

    # n = int(driver.find_element_by_css_selector('div.eUyZxf span.rua5Nb').text.strip('(').strip(')'))  # no. of participants present

    wb = load_workbook('Google_Attendance.xlsx')
    sheet = wb['Attendance Sheet']

    for name in names:
        for box in range(2, sheet.max_row+1):
            if name.text.lower() == (sheet.cell(row=box, column=column_index_from_string('A')).value.lower()):  # checks if the participant is present in the sheet
                sheet[f'B{box}'] = 'Present'  # marks present of that participant

    wb.save('Google_Attendance.xlsx')
    # print(f'No. of participants : {n-1}')  # 1 participant adds up while taking attendance
    print('Attendance taken successfully!')
